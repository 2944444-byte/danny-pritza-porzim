"""
excel_service.py
----------------
Excel helpers shared by the API:
  - read_rows(): parse an uploaded .xlsx into a list of row dicts.
  - build_workbook_bytes(): build an .xlsx from validated rows.

Uses openpyxl only (no pandas), which keeps the dependency footprint small —
important for serverless deployment (e.g. Vercel) where function size is limited.
"""

import io
from typing import Any, Dict, List

import openpyxl

from src.tables.table_handler import TableSchemaManager


def read_sheet(contents: bytes) -> tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse the bytes of an .xlsx file into (headers, records):
      - headers: the first row's column names (as-is; caller normalizes).
      - records: one {header: value} dict per data row; empty cells become ""
        and fully-empty rows are skipped.

    Returning headers separately lets callers detect column-name mismatches even
    when the sheet has headers but no data rows.
    """
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    headers = ["" if h is None else str(h) for h in header_row]

    records: List[Dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(cell is None for cell in row):
            continue
        record: Dict[str, Any] = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            record[header] = "" if value is None else value
        records.append(record)
    return headers, records


import re

# Characters that are unsafe in a download filename.
_INVALID_FILE_CHARS = re.compile(r'[\\/:*?"<>|]+')


def safe_filename(title: str | None, default: str = "phone_mappings") -> str:
    """Turn a user title into a safe `.xlsx` download filename (whole-file name)."""
    cleaned = _INVALID_FILE_CHARS.sub("", (title or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    name = cleaned or default
    return name if name.lower().endswith(".xlsx") else f"{name}.xlsx"


def build_workbook_bytes(rows: List[Dict[str, Any]], schema: TableSchemaManager) -> bytes:
    """Return the bytes of an .xlsx file containing `rows` under the schema headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phone Mappings"

    # Header row (human-friendly display names).
    ws.append(schema.display_headers)

    # Data rows, in canonical column order.
    for row in rows or []:
        ws.append([row.get(key, "") for key in schema.keys])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
